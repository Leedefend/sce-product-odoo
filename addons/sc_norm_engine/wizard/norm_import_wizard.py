import base64
import io
import re

from openpyxl import load_workbook

from odoo import _, fields, models
from odoo.exceptions import UserError


def _clean(val):
    if val is None:
        return ""
    text = str(val)
    text = text.replace("\n", "").replace("\r", "").replace("　", " ")
    return text.strip()


def _is_chapter_code(code):
    code = _clean(code)
    return bool(re.fullmatch(r"[A-Z]{2}", code))


def _detect_specialty_from_title(title):
    """从 'A 土建与装饰定额' 提取 (A, 土建与装饰)。"""
    title = _clean(title)
    if not title:
        return None, None
    title = re.sub(r"^[▲■●]", "", title).strip()
    match = re.match(r"^([A-Z])\s*(.+)$", title)
    if not match:
        return None, None
    code = match.group(1)
    name = match.group(2).replace("定额", "").strip()
    return code, name


def _to_float(val):
    val = _clean(val)
    if not val:
        return 0.0
    val = val.replace(",", "")
    try:
        return float(val)
    except Exception:
        return 0.0


class ScNormImportWizard(models.TransientModel):
    _name = "sc.norm.import.wizard"
    _description = "四川2015定额导入"

    data_file = fields.Binary("定额文件", required=True)
    filename = fields.Char("文件名")
    clear_before = fields.Boolean("导入前清空已有定额数据", default=False)
    log = fields.Text("导入日志", readonly=True)

    def action_import(self):
        self.ensure_one()
        if not self.data_file:
            raise UserError(_("请先上传定额文件！"))

        try:
            data = base64.b64decode(self.data_file)
            workbook = load_workbook(io.BytesIO(data), data_only=True)
        except Exception as exc:
            raise UserError(_("加载 Excel 失败：%s") % exc)

        sheet_names = workbook.sheetnames
        index_sheet_names = [n for n in sheet_names if n in ("专业章节", "专业章节(安装)")]

        Specialty = self.env["sc.norm.specialty"]
        Chapter = self.env["sc.norm.chapter"]
        Item = self.env["sc.norm.item"]

        if self.clear_before:
            Item.sudo().search([]).unlink()
            Chapter.sudo().search([]).unlink()
            Specialty.sudo().search([]).unlink()

        log_lines = []

        spec_index_info = self._parse_index_sheets(workbook, index_sheet_names)
        log_lines.append("识别到专业数量：%s" % len(spec_index_info))

        spec_by_code = {}
        chap_by_key = {}
        for spec_code, info in spec_index_info.items():
            name = info.get("specialty_name") or spec_code
            sheet_name = info.get("sheet_name", "")

            spec = Specialty.sudo().search([("code", "=", spec_code)], limit=1)
            if not spec:
                spec = Specialty.sudo().create(
                    {"code": spec_code, "name": name, "sheet_name": sheet_name}
                )
            else:
                spec.sudo().write({"name": name, "sheet_name": sheet_name})
            spec_by_code[spec_code] = spec
            log_lines.append(f"专业 {spec_code} - {name}")

            for seq, ch in enumerate(info.get("chapters", []), start=10):
                chap = Chapter.sudo().search(
                    [("specialty_id", "=", spec.id), ("code", "=", ch["chapter_code"])],
                    limit=1,
                )
                vals = {
                    "specialty_id": spec.id,
                    "code": ch["chapter_code"],
                    "name": ch["chapter_name"],
                    "sequence": seq,
                }
                if chap:
                    chap.sudo().write(vals)
                else:
                    chap = Chapter.sudo().create(vals)
                chap_by_key[(spec.code, ch["chapter_code"])] = chap

        imported_count = 0
        for sheet_name in sheet_names:
            clean_name = _clean(sheet_name)
            if clean_name.startswith(("封面", "项目管理", "专业章节")):
                continue
            if "定额" not in clean_name:
                continue

            spec_code = clean_name[0]
            spec = spec_by_code.get(spec_code)
            if not spec:
                spec = Specialty.sudo().search([("code", "=", spec_code)], limit=1)
                if not spec:
                    spec = Specialty.sudo().create(
                        {
                            "code": spec_code,
                            "name": clean_name[1:].replace("定额", "").strip() or spec_code,
                            "sheet_name": sheet_name,
                        }
                    )
                spec_by_code[spec_code] = spec

            ws = workbook[sheet_name]
            max_row = ws.max_row
            max_col = ws.max_column

            header = [_clean(ws.cell(row=1, column=c).value) for c in range(1, max_col + 1)]
            col_index = {name: idx + 1 for idx, name in enumerate(header) if name}

            def col(name, default=None):
                return col_index.get(name, default)

            col_no = col("序号")
            col_code = col("定额名称")
            col_name = col("项目名称")
            col_unit = col("单位")
            col_price = col("综合单价")
            col_direct = col("直接费")
            col_labor = col("人工费")
            col_material = col("材料费")
            col_machine = col("机械费")
            col_fee_rate = col("费率") or col("机械费率")
            col_misc = col("综合费")
            col_work = col("工作内容")

            if not col_code or not col_name:
                log_lines.append(f"- 工作表 {sheet_name}: 找不到定额编号/项目名称列，跳过")
                continue

            for r in range(2, max_row + 1):
                code = _clean(ws.cell(row=r, column=col_code).value)
                name = _clean(ws.cell(row=r, column=col_name).value)
                if not code or not name:
                    continue

                unit_raw = _clean(ws.cell(row=r, column=col_unit).value) if col_unit else ""
                price_total = _to_float(ws.cell(row=r, column=col_price).value) if col_price else 0.0
                cost_direct = _to_float(ws.cell(row=r, column=col_direct).value) if col_direct else 0.0
                cost_labor = _to_float(ws.cell(row=r, column=col_labor).value) if col_labor else 0.0
                cost_material = _to_float(ws.cell(row=r, column=col_material).value) if col_material else 0.0
                cost_machine = _to_float(ws.cell(row=r, column=col_machine).value) if col_machine else 0.0
                fee_rate = _to_float(ws.cell(row=r, column=col_fee_rate).value) if col_fee_rate else 0.0
                cost_misc = _to_float(ws.cell(row=r, column=col_misc).value) if col_misc else 0.0
                work_desc = _clean(ws.cell(row=r, column=col_work).value) if col_work else ""

                chapter_code = code[:2]
                chapter = chap_by_key.get((spec.code, chapter_code))

                vals = {
                    "code": code,
                    "name": name,
                    "specialty_id": spec.id,
                    "chapter_id": chapter.id if chapter else False,
                    "unit_raw": unit_raw,
                    "price_total": price_total,
                    "cost_direct": cost_direct,
                    "cost_labor": cost_labor,
                    "cost_material": cost_material,
                    "cost_machine": cost_machine,
                    "fee_rate": fee_rate,
                    "cost_misc": cost_misc,
                    "work_desc": work_desc,
                    "line_no": (
                        int(ws.cell(row=r, column=col_no).value)
                        if col_no and ws.cell(row=r, column=col_no).value
                        else r
                    ),
                }

                existing = Item.sudo().search(
                    [("specialty_id", "=", spec.id), ("code", "=", code)],
                    limit=1,
                )
                if existing:
                    existing.sudo().write(vals)
                else:
                    Item.sudo().create(vals)
                imported_count += 1

            log_lines.append(f"- 工作表 {sheet_name} ({spec.code}) 导入完成")

        log_lines.append(f"总共导入/更新定额子目：{imported_count} 条")
        self.log = "\n".join(log_lines)

        return {
            "type": "ir.actions.act_window",
            "res_model": "sc.norm.import.wizard",
            "view_mode": "form",
            "res_id": self.id,
            "target": "new",
        }

    # ================== 辅助：解析目录 sheet ==================
    def _parse_index_sheets(self, workbook, sheet_names):
        """
        返回:
        {
          "A": {
            "specialty_name": "土建与装饰",
            "chapters": [{"chapter_code": "AA", "chapter_name": "A.A 土石方工程"}, ...],
            "sheet_name": "A土建与装饰定额"
          },
          ...
        }
        """
        result = {}
        for sheet_name in sheet_names:
            ws = workbook[sheet_name]
            info = self._parse_single_index_sheet(ws)
            for code, val in info.items():
                if code not in result:
                    result[code] = val
                else:
                    result[code]["chapters"].extend(val.get("chapters", []))

        for spec_code, val in result.items():
            for name in workbook.sheetnames:
                clean = _clean(name)
                if clean.startswith(spec_code) and "定额" in clean:
                    val["sheet_name"] = name
                    break
        return result

    def _parse_single_index_sheet(self, ws):
        from collections import defaultdict

        res = defaultdict(lambda: {"specialty_name": None, "chapters": []})
        max_row = ws.max_row
        max_col = ws.max_column

        header_row = None
        chapter_col = None
        name_col = None

        for r in range(1, min(20, max_row) + 1):
            row_vals = [_clean(ws.cell(row=r, column=c).value) for c in range(1, max_col + 1)]
            if "章节" in row_vals and "名称" in row_vals:
                header_row = r
                for c, v in enumerate(row_vals, start=1):
                    if v == "章节":
                        chapter_col = c
                    elif v == "名称":
                        name_col = c
                break

        if header_row is None:
            return {}
        if chapter_col is None:
            chapter_col = 2
        if name_col is None:
            name_col = 3

        for r in range(header_row + 1, max_row + 1):
            val_chapter = _clean(ws.cell(row=r, column=chapter_col).value)
            val_name = _clean(ws.cell(row=r, column=name_col).value)

            spec_code, spec_name = _detect_specialty_from_title(val_name)
            if spec_code and spec_name:
                if not res[spec_code]["specialty_name"]:
                    res[spec_code]["specialty_name"] = spec_name

            if _is_chapter_code(val_chapter) and val_name:
                spec_code = val_chapter[0]
                res[spec_code]["chapters"].append(
                    {"chapter_code": val_chapter, "chapter_name": val_name}
                )
        return res
