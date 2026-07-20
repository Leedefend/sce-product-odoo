# -*- coding: utf-8 -*-
import base64
import csv
import io
import logging
import re

from odoo import fields, models
from odoo.exceptions import UserError
from odoo.tools import misc

from ..models.support.state_guard import raise_guard

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None


_logger = logging.getLogger(__name__)


class ProjectBoqImportWizard(models.TransientModel):
    _name = "project.boq.import.wizard"
    _description = "工程量清单导入"

    BATCH_CREATE_SIZE = 500

    UOM_ALIAS_MAP = {
        "m2": ["㎡", "m²", "平米", "平方米", "平方"],
        "m3": ["立方", "立方米"],
        "item": ["项", "项（包干）", "项(包干)", "item"],
    }

    project_id = fields.Many2one(
        "project.project",
        string="项目",
        required=True,
    )
    section_type = fields.Selection(
        [
            ("building", "建筑"),
            ("installation", "安装/机电"),
            ("decoration", "装饰"),
            ("landscape", "景观"),
            ("other", "其他"),
        ],
        string="工程类别",
        help="若文件未识别到工程类别，使用此处的默认值。",
    )
    boq_category = fields.Selection(
        [
            ("boq", "分部分项清单"),
            ("unit_measure", "单价措施清单"),
            ("total_measure", "总价措施清单"),
            ("fee", "规费"),
            ("tax", "税金"),
            ("other", "其他费用"),
        ],
        string="清单类别",
        default="boq",
        required=True,
        help="标识清单来源类别，避免分部分项与措施清单互相混淆。",
    )
    single_name = fields.Char("单项工程")
    unit_name = fields.Char("单位工程")
    source_type = fields.Selection(
        [
            ("tender", "招标清单"),
            ("contract", "合同清单"),
            ("settlement", "结算清单"),
        ],
        string="清单来源",
        default="contract",
    )
    version = fields.Char("版本号/批次", default="V1")
    clear_mode = fields.Selection(
        [
            ("append", "追加"),
            ("replace_project", "清空当前项目后导入"),
            ("replace_code", "按编码覆盖"),
        ],
        string="导入策略",
        default="append",
        required=True,
    )
    file = fields.Binary(string="导入文件", required=True)
    filename = fields.Char("文件名")
    log = fields.Text("导入日志", readonly=True)
    note = fields.Html(
        string="导入说明",
        readonly=True,
        default=lambda self: (
            "<ul>"
            "<li>同一编码在表中多次出现，将导入为多条清单行，并在工程结构中归入同一清单子目节点。</li>"
            "<li>若单位不存在，将自动规范化并创建新的计量单位。</li>"
            "<li>导入策略：追加 / 清空项目后导入 / 按编码覆盖。</li>"
            "</ul>"
        ),
    )

    # -------------------------------------------------------------------------
    # 主入口
    # -------------------------------------------------------------------------
    def action_import(self):
        self.ensure_one()
        if not self.file:
            raise UserError("请先上传导入文件。")
        if self.project_id and self.project_id.is_boq_frozen() and self.clear_mode in ("replace_project", "replace_code"):
            raise_guard(
                "P0_BOQ_FROZEN",
                f"项目[{self.project_id.display_name}]",
                "覆盖/清空导入 BOQ",
                reasons=["项目已进入结算/支付关键节点"],
                hints=["请先完成/撤销结算或付款流程后再进行覆盖导入"],
            )

        rows, created_uoms, skipped = self._parse_file()
        if not rows:
            raise UserError(
                "未找到可导入的清单数据：\n"
                "请确认文件中至少有一行同时包含名称列（清单名称/项目名称/汇总内容等）"
                "并且数量/单价/金额至少有一个为数字。"
            )

        Boq = self.env["project.boq.line"]
        created_count = 0
        updated_count = 0

        def _create_rows(vals_list):
            """按行的 boq_category 决定是否启用层级导入。"""
            if not vals_list:
                return 0

            grouped = {}
            for vals in vals_list:
                cat = vals.get("boq_category") or self.boq_category or "boq"
                grouped.setdefault(cat, []).append(vals)

            created = 0
            for cat, chunk in grouped.items():
                if cat in ("boq", "other"):
                    created += self._create_with_hierarchy(Boq, chunk)
                else:
                    created += self._batch_create(Boq, chunk)
            return created

        if self.clear_mode == "replace_project":
            # 安全检查：若已有合同清单引用本项目清单，禁止整表清空，避免外键错误
            linked_lines = self.env["construction.contract.line"].sudo().search_count(
                [("boq_line_id.project_id", "=", self.project_id.id)]
            )
            if linked_lines:
                raise UserError(
                    "当前项目的清单已被合同引用，禁止“清空项目后导入”。\n"
                    "请创建新预算版本或使用“按编码覆盖/追加”策略。"
                )
            Boq.search([("project_id", "=", self.project_id.id)]).unlink()
            created_count = _create_rows(rows)
        elif self.clear_mode == "replace_code":
            for vals in rows:
                domain = [
                    ("project_id", "=", vals["project_id"]),
                    ("code", "=", vals.get("code")),
                    ("boq_category", "=", vals.get("boq_category", False)),
                    ("source_type", "=", vals.get("source_type")),
                    ("version", "=", vals.get("version")),
                ]
                existing = Boq.search(domain, limit=1)
                if existing:
                    existing.write(vals)
                    updated_count += 1
                else:
                    Boq.create(vals)
                    created_count += 1
        else:
            created_count = _create_rows(rows)

        log_lines = []
        log_lines.append(f"成功导入 {created_count} 条，更新 {updated_count} 条。")
        if skipped:
            log_lines.append(f"跳过 {skipped} 行（空行/小计行/无数值行）。")
        if created_uoms:
            log_lines.append("自动创建计量单位：\n- " + "\n- ".join(sorted(created_uoms)))
            self.project_id.message_post(body=log_lines[-1])
        log_lines.append("如需刷新工程结构，请在项目中点击“生成工程结构”按钮。")
        self.log = "\n".join(log_lines)

        return {
            "type": "ir.actions.act_window",
            "res_model": "project.boq.line",
            "view_mode": "tree,form",
            "domain": [("project_id", "=", self.project_id.id)],
            "context": {"default_project_id": self.project_id.id},
            "target": "current",
        }

    # -------------------------------------------------------------------------
    # 文件解析入口
    # -------------------------------------------------------------------------
    def _parse_file(self):
        """Parse CSV/XLS/XLSX into vals list for project.boq.line."""
        data = base64.b64decode(self.file)
        filename = (self.filename or "").lower()
        parser = BoqParser(self)
        return parser.parse_file(data, filename)

    def _parse_csv_content(self, content):
        reader = csv.reader(io.StringIO(content))
        rows_data = list(reader)
        if not rows_data:
            raise UserError("导入文件没有数据，请检查。")
        # 头部探测：使用第一行作为表头
        headers = [str(h or "").strip() for h in rows_data[0]]
        data_rows = rows_data[1:]
        col_map = self._prepare_col_map(headers)
        rows, created_uoms, skipped = self._build_rows_from_iter(
            data_rows,
            col_map,
            strict_numeric=True,
            default_single=self.single_name,
            default_unit=self.unit_name,
            boq_category=self.boq_category,
        )
        if not rows:
            rows, created_uoms, skipped = self._build_rows_from_iter(
                data_rows,
                col_map,
                strict_numeric=False,
                default_single=self.single_name,
                default_unit=self.unit_name,
                boq_category=self.boq_category,
            )
        return rows, created_uoms, skipped

    def _parse_excel(self, data, filename):
        """解析 XLS/XLSX 为 project.boq.line 的 vals 列表。"""
        col_map_cfg = self._col_map_cfg()
        rows_all = []
        created_uoms_all = set()
        skipped_all = 0

        # ---------------- XLSX ----------------
        if filename.endswith(".xlsx"):
            if not openpyxl:
                raise UserError("服务器缺少 openpyxl，无法解析 XLSX，请安装依赖或改用 CSV。")
            book = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)

            for idx, sheet in enumerate(book.worksheets, start=1):
                title = sheet.title or ""
                # 根据 sheet 名称分类；封面/汇总等直接跳过
                sheet_type, sheet_category = self._classify_sheet_title(title)
                if sheet_type in ("cover", "summary", "other_skip"):
                    continue

                if not self._is_supported_sheet(title):
                    continue

                # 表头行数按表类型区分：总价/规费/税金等通常只有 1 行列标题
                if sheet_type in ("total_measure", "fee", "tax", "other_item"):
                    header_info = self._extract_excel_header(sheet, header_rows=1)
                else:
                    header_info = self._extract_excel_header(sheet)
                if not header_info:
                    continue
                headers, data_start_row, header_row_idx = header_info

                # 解析表头前几行的“工程名称：项目\单位【专业】”
                parsed_single, parsed_unit, parsed_major = self._parse_engineering_header_excel(
                    sheet, limit=max(5, header_row_idx)
                )

                default_single = self.single_name or parsed_single
                default_unit = self.unit_name or parsed_unit

                section_type = (
                    self.section_type
                    or self._map_major_to_section_type(parsed_major)
                    or self._guess_section_type(title)
                )
                # 根据 sheet 名推断清单类别：分部分项 / 单价措施 / 总价措施 / 规费 / 税金 / 其他项目
                category = sheet_category or self._detect_boq_category(title) or self.boq_category

                col_map = self._prepare_col_map(headers, col_map_cfg)
                data_rows = [
                    list(row)
                    for row in sheet.iter_rows(min_row=data_start_row, values_only=True)
                ]

                # 清单类 sheet 统一 strict_numeric=False，
                # 标题/分部行会先保留，由内部的小计/合计过滤逻辑做最后筛选
                if category == "other":
                    rows, skipped = self._build_rows_other(
                        data_rows,
                        sheet_index=idx,
                        sheet_name=title,
                        section_type=section_type,
                        default_single=default_single,
                        default_unit=default_unit,
                        major_name=parsed_major,
                    )
                    rows_all.extend(rows)
                    skipped_all += skipped
                    continue

                rows, created_uoms, skipped = self._build_rows_from_iter(
                    data_rows,
                    col_map,
                    section_type=section_type,
                    strict_numeric=False,
                    default_single=default_single,
                    default_unit=default_unit,
                    major_name=parsed_major,
                    sheet_index=idx,
                    sheet_name=title,
                    boq_category=category,
                )

                rows_all.extend(rows)
                created_uoms_all.update(created_uoms)
                skipped_all += skipped

            return rows_all, created_uoms_all, skipped_all

        # ---------------- XLS ----------------
        if filename.endswith(".xls"):
            if not xlrd:
                raise UserError("服务器缺少 xlrd，无法解析 XLS，请安装依赖或改用 CSV。")
            book = xlrd.open_workbook(file_contents=data)

            for idx, sheet in enumerate(book.sheets(), start=1):
                if sheet.nrows < 1:
                    continue

                title = sheet.name or ""
                sheet_type, sheet_category = self._classify_sheet_title(title)
                if sheet_type in ("cover", "summary", "other_skip"):
                    continue

                if not self._is_supported_sheet(title):
                    continue

                if sheet_type in ("total_measure", "fee", "tax", "other_item"):
                    headers, data_start_row, header_row_idx = self._extract_xls_header(sheet, header_rows=1)
                else:
                    headers, data_start_row, header_row_idx = self._extract_xls_header(sheet)
                if not headers:
                    continue

                parsed_single, parsed_unit, parsed_major = self._parse_engineering_header_xls(
                    sheet, limit=max(5, header_row_idx)
                )

                default_single = self.single_name or parsed_single
                default_unit = self.unit_name or parsed_unit

                section_type = (
                    self.section_type
                    or self._map_major_to_section_type(parsed_major)
                    or self._guess_section_type(title)
                )
                category = sheet_category or self._detect_boq_category(title) or self.boq_category

                col_map = self._prepare_col_map(headers, col_map_cfg)
                data_rows = [
                    [sheet.cell_value(r, c) for c in range(sheet.ncols)]
                    for r in range(data_start_row, sheet.nrows)
                ]

                if category == "other":
                    rows, skipped = self._build_rows_other(
                        data_rows,
                        sheet_index=idx,
                        sheet_name=title,
                        section_type=section_type,
                        default_single=default_single,
                        default_unit=default_unit,
                        major_name=parsed_major,
                    )
                    rows_all.extend(rows)
                    skipped_all += skipped
                    continue

                rows, created_uoms, skipped = self._build_rows_from_iter(
                    data_rows,
                    col_map,
                    section_type=section_type,
                    strict_numeric=False,
                    default_single=default_single,
                    default_unit=default_unit,
                    major_name=parsed_major,
                    sheet_index=idx,
                    sheet_name=title,
                    boq_category=category,
                )

                rows_all.extend(rows)
                created_uoms_all.update(created_uoms)
                skipped_all += skipped

            return rows_all, created_uoms_all, skipped_all

        # ---------------- Fallback: 按 CSV 解析 ----------------
        return self._parse_csv_bytes(data), set(), 0

    # -------------------------------------------------------------------------
    # Sheet 名称分类（核心升级点）
    # -------------------------------------------------------------------------
    @staticmethod
    def _classify_sheet_title(sheet_title):
        """
        根据 sheet 名称判断：
        - sheet_type: 业务上的表类型（boq/unit_measure/total_measure/fee/tax/other）
        - category:   写入 project.boq.line.boq_category 的值

        返回: (sheet_type, category)；都可能是 None。
        """
        title_raw = sheet_title or ""
        # 去掉空格/全角空格，全部小写方便匹配
        text = title_raw.replace(" ", "").replace("\u3000", "").lower()

        # 1) 总价措施项目清单计价表（优先匹配，避免被“措施”二字误判成单价措施）
        if "总价措施项目清单计价表" in text or "总价措施项目清单" in text:
            return "total_measure", "total_measure"

        # 2) 单价措施项目清单
        if "单价措施项目清单" in text or "单价措施" in text:
            return "unit_measure", "unit_measure"

        # 3) 分部分项工程量清单
        if "分部分项工程量清单" in text or "分部分项工程清单" in text:
            return "boq", "boq"

        # 4) 其他项目清单 / 计价汇总表
        if "其他项目清单计价汇总表" in text or "其他项目清单" in text or "其他项目" in text:
            return "other_item", "other"

        # 5) 规费/税金 等专门表（有的模板会单独拆出来）
        if "规费" in text and "清单" in text:
            return "fee", "fee"
        if "税金" in text and "清单" in text:
            return "tax", "tax"

        # 6) 其它含“措施”但没说总价/单价的，尽量保守按 total_measure 处理
        if "总价措施" in text:
            return "total_measure", "total_measure"
        if "措施项目清单" in text:
            # 如果没提“单价/总价”，按单价措施兜底
            return "unit_measure", "unit_measure"

        # 7) 实在看不出来，一律当分部分项
        return None, None
    # -------------------------------------------------------------------------
    # 表头 & 列映射
    # -------------------------------------------------------------------------
    def _col_map_cfg(self):
        return {
            "code": ["清单编码", "编码", "code"],
            "name": ["清单名称", "名称", "name", "项目名称", "汇总内容"],
            "spec": ["规格", "规格型号", "项目特征", "项目特征描述"],
            "uom_id": ["单位", "uom"],
            "quantity": ["工程量", "数量", "qty"],
            "price": ["单价", "price"],
            # “金额（元）”多见于总标题，不直接当金额列匹配
            "amount": ["合价", "合计", "amount", "金额", "金额元"],
            "cost_item_id": ["成本项", "成本科目"],
            "remark": ["备注", "说明"],
            # --- 总价措施/规费类专用（目前先读出来，后面可扩展成模型字段） ---
            "rate": ["费率", "费率(%)", "费率（%）"],
            "calc_base": ["计算基础", "计费基础"],
        }

    def _prepare_col_map(self, headers, col_map_cfg=None):
        col_map_cfg = col_map_cfg or self._col_map_cfg()
        col_map = {}
        for idx, title in enumerate(headers):
            title_norm = self._normalize_header(title)
            for field, aliases in col_map_cfg.items():
                matched = False
                for alias in aliases:
                    alias_norm = self._normalize_header(alias)
                    if (
                        title_norm == alias_norm
                        or title_norm.endswith(alias_norm)
                        or alias_norm in title_norm
                    ):
                        matched = True
                        break
                if matched and field not in col_map:
                    col_map[field] = idx
                    break
        if "name" not in col_map:
            # 兜底：首列作为名称
            if headers:
                col_map["name"] = 0
            else:
                raise UserError("模板中至少需要包含 “清单名称” 列。")
        # 若识别到工程量列，按相对位置推断单价/合价（常见 F1-1 结构：工程量右一列=单价，右二列=合价）
        qty_idx = col_map.get("quantity")
        if qty_idx is not None:
            if "price" not in col_map and qty_idx + 1 < len(headers):
                col_map["price"] = qty_idx + 1
            if "amount" not in col_map and qty_idx + 2 < len(headers):
                col_map["amount"] = qty_idx + 2
        return col_map

    def _find_header_in_sheet(self, row_iter):
        """Deprecated: use _extract_excel_header/_extract_xls_header."""
        return [], 0

    # ----------------- Excel helpers -----------------
    def _parse_engineering_header_excel(self, sheet, limit=5):
        """
        解析表头中的“工程名称：项目\\单位【专业】”
        返回 (single_name, unit_name, major_name)
        """
        for row in sheet.iter_rows(min_row=1, max_row=limit, values_only=True):
            for val in row:
                if not val:
                    continue
                text = str(val)
                if "工程名称" not in text:
                    continue
                text = text.split("工程名称", 1)[-1]
                text = text.lstrip("：:").strip()
                parts = text.split("\\")
                single = parts[0].strip() if parts else ""
                tail = parts[1].strip() if len(parts) >= 2 else ""
                unit = tail
                major = ""
                if "【" in tail and "】" in tail:
                    before = tail.split("【", 1)[0]
                    inner = tail.split("【", 1)[1].split("】", 1)[0]
                    unit = before.strip()
                    major = inner.strip()
                return single, unit, major
        return "", "", ""

    def _parse_engineering_header_xls(self, sheet, limit=5):
        for r in range(min(limit, sheet.nrows)):
            for c in range(sheet.ncols):
                val = sheet.cell_value(r, c)
                if not val:
                    continue
                text = str(val)
                if "工程名称" not in text:
                    continue
                text = text.split("工程名称", 1)[-1]
                text = text.lstrip("：:").strip()
                parts = text.split("\\")
                single = parts[0].strip() if parts else ""
                tail = parts[1].strip() if len(parts) >= 2 else ""
                unit = tail
                major = ""
                if "【" in tail and "】" in tail:
                    before = tail.split("【", 1)[0]
                    inner = tail.split("【", 1)[1].split("】", 1)[0]
                    unit = before.strip()
                    major = inner.strip()
                return single, unit, major
        return "", "", ""

    def _extract_excel_header(self, sheet, header_rows=2, scan_rows=8):
        """处理多行表头+合并单元格，返回(扁平列名列表, 数据起始行号, 识别到的表头行号)"""
        merge_map = {}
        try:
            merge_ranges = sheet.merged_cells
        except Exception:
            merge_ranges = None
        if merge_ranges:
            ranges = getattr(merge_ranges, "ranges", merge_ranges)
            try:
                for m in ranges:
                    min_row, min_col, max_row, max_col = m.min_row, m.min_col, m.max_row, m.max_col
                    for r in range(min_row, max_row + 1):
                        for c in range(min_col, max_col + 1):
                            merge_map[(r, c)] = (min_row, min_col)
            except Exception:
                merge_map = {}

        def cell_val(r, c):
            key = merge_map.get((r, c))
            if key:
                r, c = key
            return sheet.cell(row=r, column=c).value

        max_col = sheet.max_column or 0
        header_row_idx = 0
        best_hits = 0
        keywords = [
            "编码",
            "项目编码",
            "清单编码",
            "特征",
            "工程量",
            "综合单价",
            "合价",
            "计量单位",
            # 其他项目清单/计价汇总表常见列
            "项目名称",
            "金额",
            "金额(元)",
            "金额（元）",
            "备注",
            "序号",
        ]
        for idx in range(1, min(scan_rows, sheet.max_row or 0) + 1):
            row_vals = [str(cell_val(idx, c) or "").strip() for c in range(1, max_col + 1)]
            hits = sum(1 for v in row_vals if any(k in v for k in keywords))
            if hits > best_hits:
                best_hits = hits
                header_row_idx = idx
        if not header_row_idx:
            return None

        header_rows_vals = []
        for r in range(header_row_idx, min(header_row_idx + header_rows, (sheet.max_row or 0) + 1)):
            row_vals = []
            for c in range(1, max_col + 1):
                row_vals.append(str(cell_val(r, c) or "").strip())
            header_rows_vals.append(row_vals)

        # 纵向拼接列名
        flat_headers = []
        for c in range(max_col):
            parts = []
            for r in range(len(header_rows_vals)):
                v = header_rows_vals[r][c]
                if v:
                    parts.append(v)
            flat_headers.append(" - ".join(parts) if parts else "")

        data_start = header_row_idx + header_rows
        return flat_headers, data_start, header_row_idx

    def _extract_xls_header(self, sheet, header_rows=2, scan_rows=8):
        max_col = sheet.ncols
        header_row_idx = 0
        best_hits = 0
        keywords = [
            "编码",
            "项目编码",
            "清单编码",
            "特征",
            "工程量",
            "综合单价",
            "合价",
            "计量单位",
            # 其他项目清单/计价汇总表常见列
            "项目名称",
            "金额",
            "金额(元)",
            "金额（元）",
            "备注",
            "序号",
        ]
        for idx in range(min(scan_rows, sheet.nrows)):
            row_vals = [str(sheet.cell_value(idx, c) or "").strip() for c in range(max_col)]
            hits = sum(1 for v in row_vals if any(k in v for k in keywords))
            if hits > best_hits:
                best_hits = hits
                header_row_idx = idx
        if max_col == 0:
            return None, 0, 0
        header_rows_vals = []
        for r in range(header_row_idx, min(header_row_idx + header_rows, sheet.nrows)):
            row_vals = [str(sheet.cell_value(r, c) or "").strip() for c in range(max_col)]
            header_rows_vals.append(row_vals)
        flat_headers = []
        for c in range(max_col):
            parts = [row_vals[c] for row_vals in header_rows_vals if row_vals[c]]
            flat_headers.append(" - ".join(parts) if parts else "")
        data_start = header_row_idx + header_rows
        return flat_headers, data_start, header_row_idx

    # -------------------------------------------------------------------------
    # 行解析 & 单位/成本项匹配
    # -------------------------------------------------------------------------
    def _build_rows_from_iter(
        self,
        row_iter,
        col_map,
        section_type=None,
        strict_numeric=True,
        default_single=None,
        default_unit=None,
        major_name=None,
        sheet_index=None,
        sheet_name=None,
        boq_category=None,
    ):
        Uom = self.env["uom.uom"]
        Dict, dict_domain_key = self._get_dictionary_model()

        rows = []
        uom_cache = {}
        cost_item_cache = {}
        created_uoms = set()
        skipped_rows = 0
        current_division = None

        def _default_uom_category():
            """选用通用“单位”类别，若缺失则取任一类别兜底。"""
            category = self.env.ref("uom.product_uom_categ_unit", raise_if_not_found=False)
            if not category:
                category = self.env["uom.category"].search([], limit=1)
            return category

        for row in row_iter:
            
             #小工具：按字段名取这一行对应列的值
            def get(field):
                idx = col_map.get(field)
                if idx is None or idx >= len(row):
                    return ""
                return row[idx] if not isinstance(row, dict) else row.get(idx)

            name = str(get("name") or "").strip()
            code = str(get("code") or "").strip()
            if not (name or code):
                skipped_rows += 1
                continue

            eff_boq_category = boq_category or self.boq_category

            # 逻辑上的“项目编码”：只保留阿拉伯数字，过滤掉序号/符号
            logical_code_digits = re.sub(r"\D", "", code or "")

            # 非分部分项清单：缺少编码直接跳过，避免必填校验失败
            if eff_boq_category != "boq" and not code:
                skipped_rows += 1
                continue

            # 分部分项清单：项目编码空、名称有值（且非数字） -> 记录当前分部，跳过本行
            if (
                eff_boq_category == "boq"
                and not code
                and name
                and not self._is_number(name)
            ):
                # 小计/合计类行作为分部标题会污染分部名称，直接跳过
                lower_name = name.lower()
                if any(key in lower_name for key in ["合计", "小计", "本页", "本表"]):
                    skipped_rows += 1
                    continue
                current_division = name
                continue

            vals = {
                "project_id": self.project_id.id,
                "name": name,
                "section_type": self.section_type or section_type or False,
                "single_name": default_single or self.single_name or False,
                "unit_name": default_unit or self.unit_name or False,
                "major_name": major_name or False,
                "sheet_index": sheet_index,
                "sheet_name": sheet_name,
                "source_type": self.source_type,
                "version": self.version,
                "boq_category": eff_boq_category or "boq",
                # 默认认为都是清单项，由 _create_with_hierarchy 根据层级修正。
                "line_type": "item",
            }
            if section_type and not vals["section_type"]:
                vals["section_type"] = section_type

            spec = str(get("spec") or "").strip()
            remark = str(get("remark") or "").strip()
            qty = get("quantity")
            price = get("price")
            amount_val = get("amount")
            rate_val = get("rate")
            calc_base = str(get("calc_base") or "").strip()

            # ---- 总价措施 / 规费 / 税金 专用规则 ----
            if eff_boq_category in ("total_measure", "fee", "tax"):
                lower_name = (name or "").lower()
                # 1) 合计/本页/本表 行直接跳过
                if any(k in lower_name for k in ("合计", "本页", "本表")):
                    skipped_rows += 1
                    continue

                # 2) 子目行：逻辑上无有效项目编码 + 有金额，当前版本不导入，避免重复计入总价
                #    例如 ① / ② / ③ / ④ 这些序号会被视为无效编码
                if not logical_code_digits and self._is_number(amount_val):
                    continue

                # 3) 只有金额，没有工程量/单价 → 用金额补齐 qty/price
                if (not self._is_number(qty)) and (not self._is_number(price)) and self._is_number(amount_val):
                    qty = 1.0
                    price = amount_val

                # 4) 金额型费用行统一视为清单项
                vals["line_type"] = "item"

            if code:
                vals["code"] = code
            if spec:
                vals["spec"] = spec

            # 记录当前分部名称，便于 WBS 直接使用（无需再从 remark 里解析）。
            if current_division:
                vals["division_name"] = current_division
            if remark or current_division:
                prefix = f"[分部]{current_division}" if current_division else ""
                vals["remark"] = f"{prefix} {remark}".strip() if (prefix or remark) else False

            vals["quantity"] = self._to_float(qty)
            vals["price"] = self._to_float(price)
            # 注意：amount 字段是 compute+store，直接写值会被覆盖；
            # 这里仍然填入，便于诊断和自定义逻辑读取。
            vals["amount"] = self._to_float(amount_val)

            # 若数量/单价/合价均为0，则视为标题/小计行跳过
            if strict_numeric:
                if not any(
                    [
                        self._is_number(qty),
                        self._is_number(price),
                        self._is_number(amount_val),
                    ]
                ):
                    skipped_rows += 1
                    continue

            # 常见小计/合计行过滤
            lower_name = (name or "").lower()
            if any(key in lower_name for key in ["合计", "小计", "本页", "本表"]):
                skipped_rows += 1
                continue

            # ===== 计量单位处理 =====
            uom = False
            uom_name = str(get("uom_id") or "").strip()
            if uom_name:
                norm_name = self._normalize_uom_name(uom_name)
                canonical = self._canonical_uom(norm_name)
                search_key = canonical or norm_name or uom_name

                uom = uom_cache.get(search_key)
                create_name = None

                if uom is None:
                    # 先按规范名找
                    uom = Uom.search([("name", "=", search_key)], limit=1)
                    # 再按原始名兜底
                    if not uom and uom_name != search_key:
                        uom = Uom.search([("name", "=", uom_name)], limit=1)

                    if not uom:
                        category = _default_uom_category()
                        if not category:
                            raise UserError(
                                "未找到计量单位类别，无法自动创建单位，请先在系统中创建一个计量单位类别。"
                            )
                        create_name = search_key
                        ref_uom = Uom.search(
                            [
                                ("category_id", "=", category.id),
                                ("uom_type", "=", "reference"),
                            ],
                            limit=1,
                        )
                        uom_vals = {
                            "name": create_name,
                            "category_id": category.id,
                            "factor": 1.0,
                            "factor_inv": 1.0,
                            "rounding": 0.0001,
                            "active": True,
                        }
                        # 如果类别已有参照单位，则新建等效单位用 smaller 并保持 factor=1
                        if ref_uom:
                            uom_vals["uom_type"] = "smaller"
                            uom_vals["factor"] = 1.0
                            uom_vals["factor_inv"] = 1.0
                        else:
                            uom_vals["uom_type"] = "reference"
                        uom = Uom.create(uom_vals)
                        if create_name:
                            created_uoms.add(create_name)

                if uom:
                    uom_cache[search_key] = uom

            # 若仍未找到单位，使用“单位”类别的参照单位兜底，避免必填校验失败
            if not uom:
                category = _default_uom_category()
                if category:
                    uom = Uom.search(
                        [
                            ("category_id", "=", category.id),
                            ("uom_type", "=", "reference"),
                        ],
                        limit=1,
                    )
                    if not uom:
                        # 创建一个通用参照单位“项”作为兜底
                        uom = Uom.create(
                            {
                                "name": "项",
                                "category_id": category.id,
                                "uom_type": "reference",
                                "factor": 1.0,
                                "factor_inv": 1.0,
                                "rounding": 0.0001,
                                "active": True,
                            }
                        )
                        created_uoms.add("项")

            vals["uom_id"] = uom.id if uom else False

            # ===== 成本项字典匹配 =====
            cost_item_name = str(get("cost_item_id") or "").strip()
            if cost_item_name and Dict:
                cost_item = cost_item_cache.get(cost_item_name)
                if cost_item is None:
                    if isinstance(dict_domain_key, (list, tuple)):
                        domain = list(dict_domain_key)
                    else:
                        domain = [(dict_domain_key, "=", "cost_item")]
                    domain.append(("name", "=", cost_item_name))
                    cost_item = Dict.search(domain, limit=1)
                    cost_item_cache[cost_item_name] = cost_item
                vals["cost_item_id"] = cost_item.id or False

            # --- 总价措施 / 单价措施等“非分部分项”表的分部兜底 ---
            # 这些表本身没有“分部标题行”，为了避免 division_name=False 出现在分组视图里，
            # 这里按清单类别统一给一个可读的分部名称。
            if not vals.get("division_name"):
                if boq_category in ("total_measure",):
                    # 总价措施项目清单
                    vals["division_name"] = "总价措施项目"
                elif boq_category in ("unit_measure",):
                    # 单价措施项目清单
                    vals["division_name"] = "单价措施项目"
                elif boq_category in ("fee", "tax"):
                    vals["division_name"] = "规费及税金"

            rows.append(vals)

        return rows, created_uoms, skipped_rows

    # -------------------------------------------------------------------------
    # 其他项目清单（专用解析）
    # -------------------------------------------------------------------------
    def _build_rows_other(
        self,
        data_rows,
        sheet_index=None,
        sheet_name=None,
        section_type=None,
        default_single=None,
        default_unit=None,
        major_name=None,
    ):
        """
        解析《其他项目清单与计价汇总表》：
        - A 列：序号/层级编码（1 / 2 / 2.1 / 3 / 合计）
        - B 列：项目名称
        - C 列：金额（无数量/单价）
        """
        rows = []
        skipped = 0

        Uom = self.env["uom.uom"]

        def _default_uom():
            category = self.env.ref("uom.product_uom_categ_unit", raise_if_not_found=False)
            if not category:
                category = self.env["uom.category"].search([], limit=1)
            ref_uom = False
            if category:
                ref_uom = Uom.search(
                    [("category_id", "=", category.id), ("uom_type", "=", "reference")], limit=1
                )
                if not ref_uom:
                    ref_uom = Uom.create(
                        {
                            "name": "项",
                            "category_id": category.id,
                            "uom_type": "reference",
                            "factor": 1.0,
                            "factor_inv": 1.0,
                            "rounding": 0.0001,
                            "active": True,
                        }
                    )
            return ref_uom

        default_uom = _default_uom()

        for row in data_rows:
            code = str((row[0] if len(row) > 0 else "") or "").strip()
            name = str((row[1] if len(row) > 1 else "") or "").strip()
            amount_raw = row[2] if len(row) > 2 else ""

            # 空行直接跳过
            if not code and not name:
                skipped += 1
                continue

            # 合计/总计行（写在序号或项目名称里）一律跳过
            label = f"{code}{name}".replace(" ", "").replace("　", "")
            if label in ("合计", "本表合计", "本页合计", "总计", "台计"):
                skipped += 1
                continue

            line_type, level = self._parse_other_line_level(code)
            if not line_type or not level:
                # 兜底：当作一级标题
                line_type = "group"
                level = 1

            amount = self._to_float(amount_raw)
            qty = 1.0
            price = amount

            vals = {
                "project_id": self.project_id.id,
                "name": name,
                "code": code,
                "quantity": qty,
                "price": price,
                "amount": amount,
                "boq_category": "other",
                "division_name": "其他项目",
                "line_type": line_type,
                "sheet_index": sheet_index,
                "sheet_name": sheet_name,
                "section_type": self.section_type or section_type or False,
                "single_name": default_single or self.single_name or False,
                "unit_name": default_unit or self.unit_name or False,
                "major_name": major_name or False,
                "source_type": self.source_type,
                "version": self.version,
            }
            if default_uom:
                vals["uom_id"] = default_uom.id
            rows.append(vals)

        return rows, skipped

    # -------------------------------------------------------------------------
    # 字符串/数值工具
    # -------------------------------------------------------------------------
    def _read_as_csv(self, data_bytes):
        """Return CSV string from raw bytes."""
        return self._parse_csv_bytes(data_bytes)

    def _parse_csv_bytes(self, data_bytes):
        """Try utf-8, then gbk."""
        for encoding in ("utf-8", "gbk"):
            try:
                return data_bytes.decode(encoding)
            except Exception:
                continue
        raise UserError("无法解码导入文件，请确认使用 UTF-8 或 GBK 编码。")

    @staticmethod
    def _guess_section_type(sheet_title):
        title = (sheet_title or "").lower()
        mapping = {
            "build": "building",
            "建筑": "building",
            "机电": "installation",
            "安装": "installation",
            "elect": "installation",
            "装饰": "decoration",
            "decoration": "decoration",
            "景观": "landscape",
            "landscape": "landscape",
        }
        for key, val in mapping.items():
            if key in title:
                return val
        return False

    def _is_supported_sheet(self, title):
        """
        只要能识别出 sheet_type，就认为是“清单相关的有效 sheet”，其他一律跳过
        （封面、汇总表、投标总说明之类不会被读取）。
        """
        sheet_type, category = ProjectBoqImportWizard._classify_sheet_title(title or "")
        return bool(sheet_type)

    @staticmethod
    def _map_major_to_section_type(major_name):
        """根据专业名称映射工程类别（section_type）"""
        text = (major_name or "").lower()
        mapping = {
            "装饰": "decoration",
            "装修": "decoration",
            "建筑": "building",
            "土建": "building",
            "给排水": "installation",
            "暖通": "installation",
            "电气": "installation",
            "强电": "installation",
            "弱电": "installation",
            "机电": "installation",
            "消防": "installation",
            "安装": "installation",
            "景观": "landscape",
        }
        for key, val in mapping.items():
            if key in text:
                return val
        return False

    # 其他项目清单：根据序号判断层级与行类型
    @staticmethod
    def _parse_other_line_level(code):
        """
        返回 (line_type, level)：
        - 纯数字：一级 group
        - 数字.数字：二级 item
        - 合计：None
        """
        text = (code or "").strip()
        if text in ("合计", "合 计", "台计"):
            return None, None
        if not text:
            return None, None
        if text.isdigit():
            return "group", 1
        if re.match(r"^\d+\.\d+$", text):
            return "item", 2
        return None, None

    @staticmethod
    def _detect_boq_category(sheet_title):
        """
        根据 sheet 名推断清单类别：分部分项/单价措施/总价措施/规费/税金/其他项目。
        实现上复用 _classify_sheet_title 的逻辑。
        """
        sheet_type, category = ProjectBoqImportWizard._classify_sheet_title(sheet_title or "")
        if category:
            return category
        # 没识别到就按分部分项兜底
        return "boq"

    @staticmethod
    def _normalize_header(title):
        text = str(title or "").strip()
        text = re.sub(r"\s+", "", text)
        return text.lower()

    @staticmethod
    def _is_number(val):
        try:
            if isinstance(val, str):
                cleaned = ProjectBoqImportWizard._clean_number_str(val)
                if cleaned in ("", "-", "--"):
                    return False
                float(cleaned)
            else:
                float(val)
            return True
        except Exception:
            return False

    @staticmethod
    def _to_float(val):
        try:
            if isinstance(val, str):
                cleaned = ProjectBoqImportWizard._clean_number_str(val)
                return float(cleaned or 0.0)
            return float(val or 0.0)
        except Exception:
            return 0.0

    def _get_dictionary_model(self):
        """返回可用的字典模型及类型字段键。"""
        dict_model = "project.dictionary" if "project.dictionary" in self.env.registry else "sc.dictionary"
        if dict_model not in self.env.registry:
            return None, None
        Dict = self.env[dict_model]
        fields_map = Dict._fields
        if "type" in fields_map:
            return Dict, "type"
        if "type_id" in fields_map:
            return Dict, "type_id.code"
        return Dict, "type"

    @staticmethod
    def _clean_number_str(text):
        """Remove common thousand separators and spaces."""
        cleaned = str(text or "")
        cleaned = cleaned.replace(",", "").replace("，", "").replace(" ", "").strip()
        return cleaned

    # --- UoM helpers ---
    def _normalize_uom_name(self, name):
        """基本规范化：去空格、全角转半角、小写。"""
        text = misc.ustr(name or "").strip()
        text = re.sub(r"\s+", "", text)
        res = []
        for ch in text:
            code = ord(ch)
            if 0xFF01 <= code <= 0xFF5E:
                code -= 0xfee0
                ch = chr(code)
            res.append(ch)
        return "".join(res).lower()

    def _canonical_uom(self, norm_name):
        """根据别名映射返回规范名，否则返回原名。"""
        for main, aliases in self.UOM_ALIAS_MAP.items():
            if norm_name == main:
                return main
            for alias in aliases:
                if norm_name == self._normalize_uom_name(alias):
                    return main
        return norm_name

    # -------------------------------------------------------------------------
    # 批量创建 & 层级构建
    # -------------------------------------------------------------------------
    def _batch_create(self, model, vals_list):
        """批量创建，避免一次性巨大列表占用内存/锁时间过长。"""
        if not vals_list:
            return 0
        size = self.BATCH_CREATE_SIZE or 500
        created = 0
        for start in range(0, len(vals_list), size):
            chunk = vals_list[start : start + size]
            model.create(chunk)
            created += len(chunk)
        return created

    def _create_with_hierarchy(self, model, vals_list):
        """
        批量创建 + 根据编码/上下文推断层级，写入 parent_id + line_type。
        仅在分部分项清单（boq_category='boq'）中使用。

        思路：
        - 保持创建顺序（vals_list 的顺序），避免打乱 Excel 原始行序；
        - 以 (project, section_type, single_name, unit_name, sheet_index) 为分段 key，
          每一段单独维护一个“层级栈”（stack: level -> record）；
        - 根据编码/名称/是否有数量，调用 _classify_line 得到 (line_type, level)；
        - level=0 无 parent，level>0 时 parent = stack[level-1]；
        - 最后将当前记录放入 stack[level]，供后面的行作为下级挂接。
        """
        if not vals_list:
            return 0

        # 一次性创建所有记录（保持 vals_list 顺序）
        records = model.create(vals_list)

        # 为了保证层级构建稳定，以 sheet_index / id 排个序
        ordered_records = sorted(
            records,
            key=lambda r: (
                r.project_id.id or 0,
                r.section_type or "",
                r.single_name or "",
                r.unit_name or "",
                r.sheet_index or 0,
                r.id,
            ),
        )

        current_key = None
        stack = {}  # level(int) -> record

        for rec in ordered_records:
            key = (
                rec.project_id.id,
                rec.section_type or "",
                rec.single_name or "",
                rec.unit_name or "",
                rec.sheet_index or 0,
            )
            # 换了 sheet / 单项工程 / 单位工程：重置层级栈
            if key != current_key:
                stack = {}
                current_key = key

            # 其他项目清单：使用专用层级规则（code 决定 level）
            if rec.boq_category == "other":
                o_line_type, o_level = ProjectBoqImportWizard._parse_other_line_level(rec.code)
                if not o_line_type or o_level is None:
                    o_line_type, o_level = "group", 1
                rec.line_type = o_line_type
                if o_level <= 0:
                    rec.parent_id = False
                else:
                    parent = stack.get(o_level - 1)
                    rec.parent_id = parent.id if parent else False
                stack[o_level] = rec
                continue

            line_type, level = self._classify_line(
                code=rec.code,
                name=rec.name,
                qty=rec.quantity,
                price=rec.price,
                amount=rec.amount,
                boq_category=rec.boq_category,
            )

            # 写入行类型
            rec.line_type = line_type

            # 处理 parent_id
            if level <= 0:
                rec.parent_id = False
            else:
                parent = stack.get(level - 1)
                rec.parent_id = parent.id if parent else False

            # 记录当前层级最近一行，供后面的子级挂接。
            stack[level] = rec

        return len(records)

    @staticmethod
    def _classify_line(code, name, qty, price, amount, boq_category):
        """
        读一行，判断：
        - line_type: major / division / group / item
        - level: 0,1,2,3  对应 章 / 分部 / 小结 / 清单项

        规则分两块：
        A) 总价措施 / 规费 / 税金 等“金额型费用表”
        B) 普通分部分项清单（编码驱动层级）
        """
        code = (code or "").strip()
        name = (name or "").strip()
        lname = name.lower()

        # 是否有“数量/单价/金额”数值
        has_numeric = any(
            ProjectBoqImportWizard._is_number(v)
            for v in (qty, price, amount)
        )

        # ---------- A) 总价措施 / 规费 / 税金 ----------
        if boq_category in ("total_measure", "fee", "tax"):
            # 典型结构：第一行有项目编码，下面若干行无编码，表示该费用的组成明细
            #   1  安全文明施工费  code=0411..., amount 有值
            #   1.1 环境保护费    code 为空，费率/金额有值
            #   1.2 文明施工费    ...
            #
            # 策略：
            # - 只要这一行有 code → 视为“费用汇总行”（group, level 1）
            # - 同一表中，随后无 code 的行 → 视为该费用下的明细项（item, level 2）
            # - “合计/本表合计”等仍按通用小计过滤逻辑跳过（在别处已经处理）
            if code:
                return "group", 1
            else:
                # 无编码，但有金额/费率 → 明细项
                if has_numeric:
                    return "item", 2
                # 实在啥都没有，就当成标题行（几乎不会出现）
                return "group", 1

        # ---------- B) 普通分部分项清单：原有逻辑 ----------
        # 特殊前缀：MAJ-xxx / DIV-xxx
        if code.startswith("MAJ-") or code.startswith("MAJ"):
            return "major", 0
        if code.startswith("DIV-") or code.startswith("DIV"):
            return "division", 1

        # 纯数字/带点编码
        pure = code.replace(".", "")
        if pure.isdigit():
            length = len(pure)
            if length <= 2:
                return "major", 0
            if length <= 4:
                return "division", 1
            if length <= 6:
                return "group", 2
            # 一般 8~12 位都是具体清单项目码
            return "item", 3

        # 无编码 & 无数值：标题/汇总
        if not code and not has_numeric:
            # 合计/小计 视为小结行
            if any(k in lname for k in ("合计", "小计", "本页", "本表")):
                return "group", 2
            # 名称里包含“工程/专业”等，视为分部工程
            if any(k in name for k in ("工程", "专业", "道路", "市政", "桥梁", "绿化")):
                return "division", 1
            # 其他标题，暂当 group
            return "group", 2

        # 有数值，但是编码非数字（比如 “0401090024-01” 的特殊写法）
        if has_numeric:
            return "item", 3

        # 兜底：实在分不清的都当清单项，不破坏数据
        return "item", 3


# -------------------------------------------------------------------------
# 层级构建器（封装栈操作，保持现有层级算法）
# -------------------------------------------------------------------------
class HierarchyBuilder:
    def __init__(self):
        self.stack = {}

    def reset(self):
        self.stack = {}

    def place(self, rec, level):
        parent = self.stack.get(level - 1)
        rec.parent_id = parent.id if parent else False
        self.stack[level] = rec

    def refresh_parent_path(self, records):
        """统一刷新 parent_path；失败时保持导入流程不中断。"""
        try:
            records._parent_store_compute()
        except Exception:
            _logger.debug("Unable to refresh BOQ hierarchy parent path.", exc_info=True)

    def heal_hierarchy(self, records):
        """
        层级连续性修复扩展点。
        当前保持导入行为稳定，可用于统一调整 level/parent_id/display_order。
        """
        return records


# -------------------------------------------------------------------------
# 导入解析适配层（行为保持不变）
# -------------------------------------------------------------------------
class RowParser:
    """行解析适配器；可按清单类别扩展并保持默认行为稳定。"""

    def __init__(self, wizard):
        self.wizard = wizard

    def parse_row(self, raw_row, col_map):
        """返回原始行，供类别化解析扩展使用。"""
        return raw_row


class BoqParser:
    """
    导入解析适配层。
    当前仍委托原有 _parse_excel/_build_rows_from_iter，承担结构封装与章节池收集，
    不改变导入业务行为。
    """

    def __init__(self, wizard):
        self.wizard = wizard
        self.row_parser = RowParser(wizard)
        # 章节池：收集标题/章节文本，仅收集候选，不做层级推断。
        self.chapter_pool = []

    def parse_file(self, data, filename):
        """按文件类型分发，返回 rows/created_uoms/skipped。"""
        fname = (filename or "").lower()
        if fname.endswith((".xlsx", ".xls")):
            return self.wizard._parse_excel(data, fname)
        # CSV 默认解析
        content = self.wizard._read_as_csv(data)
        return self.wizard._parse_csv_content(content)

    def parse_sheet(self, sheet, sheet_index):
        """
        保留工作表解析扩展点，当前由 wizard._parse_excel 处理。
        预解析合并单元格标题区并收集章节池（仅收集，不推断）。
        """
        titles = self.parse_merged_title_area(sheet)
        if titles:
            self.chapter_pool.extend(titles)
        return None

    def parse_rows(self, data_rows):
        """保留行解析扩展点，当前由 wizard._build_rows_from_iter 处理。"""
        return None

    # ------------------ 章节/标题预解析 ------------------
    def parse_merged_title_area(self, sheet, max_rows=5):
        """
        简单读取前几行合并单元格的非空文本，作为章节池候选。
        仅收集文本，不做层级推断。
        """
        titles = []
        try:
            merge_ranges = sheet.merged_cells
        except Exception:
            merge_ranges = None
        ranges = getattr(merge_ranges, "ranges", merge_ranges) or []
        seen = set()
        for m in ranges:
            try:
                min_row, min_col, max_row, max_col = m.min_row, m.min_col, m.max_row, m.max_col
            except Exception:
                continue
            if min_row > max_rows:
                continue
            try:
                val = sheet.cell(row=min_row, column=min_col).value
            except Exception:
                val = None
            text = str(val or "").strip()
            if text and text not in seen:
                titles.append(text)
                seen.add(text)
        # 兜底：再扫一遍前 max_rows 行的非空单元格，补充章节文本
        for r in range(1, max_rows + 1):
            row_vals = []
            try:
                row_vals = [str(sheet.cell(row=r, column=c).value or "").strip() for c in range(1, (sheet.max_column or 0) + 1)]
            except Exception:
                _logger.debug("Unable to scan BOQ merged title row.", exc_info=True)
            for v in row_vals:
                if v and v not in seen:
                    titles.append(v)
                    seen.add(v)
        return titles
