# -*- coding: utf-8 -*-
import base64
import io
import json
import re

from odoo import api, fields, models
from odoo.exceptions import UserError

try:
    import openpyxl
except ImportError:
    openpyxl = None

from ..models.support.quota_spec import QUOTA_SPEC


def _clean(val):
    if val is None:
        return ""
    text = str(val)
    return text.replace("\n", "").replace("\r", "").replace("　", " ").strip()


def _to_float(val):
    val = _clean(val)
    if not val:
        return 0.0
    val = val.replace(",", "")
    try:
        return float(val)
    except Exception:
        return 0.0


class QuotaImportWizard(models.TransientModel):
    _name = "quota.import.wizard"
    _description = "四川定额导入向导"

    file_data = fields.Binary("定额文件", required=True)
    filename = fields.Char("文件名")
    sheet_name = fields.Char("仅导入指定工作表（可选）")
    header_row = fields.Integer("表头行号", default=1)
    log = fields.Text("导入日志", readonly=True)

    # ========== 基础工具 ==========
    def _load_workbook(self):
        if not self.file_data:
            raise UserError("请先上传定额文件。")
        if not openpyxl:
            raise UserError("服务器缺少 openpyxl 依赖，请先安装。")
        data = base64.b64decode(self.file_data)
        try:
            return openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        except Exception as exc:
            raise UserError(f"无法读取定额文件：{exc}")

    @staticmethod
    def _val(sheet, row, col):
        return sheet.cell(row=row, column=col).value if col else None

    @staticmethod
    def _str(sheet, row, col):
        v = sheet.cell(row=row, column=col).value if col else ""
        return _clean(v)

    def _ensure_dict(self, type_code, name, code=None, parent=None):
        Dic = self.env["project.dictionary"].sudo()
        domain = [("type", "=", type_code), ("name", "=", name)]
        if parent:
            domain.append(("parent_id", "=", parent.id))
        rec = Dic.search(domain, limit=1)
        if rec:
            if code and not rec.code:
                rec.code = code
            return rec
        vals = {
            "type": type_code,
            "name": name,
            "code": code,
            "parent_id": parent.id if parent else False,
        }
        return Dic.create(vals)

    def _match_sheets(self, wb):
        """
        返回 [(sheet, spec)]，支持模糊匹配 sheet 名并给出清晰报错。
        - 先做去空格匹配；
        - 支持 sheet 名以配置项开头（或反向）作为模糊匹配。
        """

        def _match_key(sheet_title):
            st = sheet_title.strip()
            for key in QUOTA_SPEC.keys():
                kt = key.strip()
                if st == kt:
                    return key
                if st.startswith(kt):
                    return key
                if kt.startswith(st):
                    return key
            return None

        pairs = []
        if self.sheet_name:
            if self.sheet_name not in wb.sheetnames:
                raise UserError(
                    "工作簿中不存在名为“%s”的工作表。\n\n当前工作簿包含：\n- %s"
                    % (self.sheet_name, "\n- ".join(wb.sheetnames))
                )
            spec_key = _match_key(self.sheet_name)
            if not spec_key:
                raise UserError(
                    "工作表“%s”未在定额配置中找到匹配项。\n\n定额配置键列表：\n- %s"
                    % (self.sheet_name, "\n- ".join(QUOTA_SPEC.keys()))
                )
            pairs.append((wb[self.sheet_name], QUOTA_SPEC[spec_key]))
        else:
            for real_title in wb.sheetnames:
                spec_key = _match_key(real_title)
                if spec_key:
                    pairs.append((wb[real_title], QUOTA_SPEC[spec_key]))

        if not pairs:
            raise UserError(
                "当前工作簿没有任何工作表能与定额配置匹配。\n\n"
                "工作簿工作表列表：\n- %s\n\n定额配置键列表：\n- %s"
                % ("\n- ".join(wb.sheetnames), "\n- ".join(QUOTA_SPEC.keys()))
            )
        return pairs

    # ========== 导入主流程 ==========
    def action_import(self):
        self.ensure_one()
        wb = self._load_workbook()
        Dic = self.env["project.dictionary"].sudo()
        Uom = self.env["uom.uom"].sudo()

        log_lines = []
        total_created = total_updated = 0

        for sheet, spec in self._match_sheets(wb):
            # 专业节点
            discipline = self._ensure_dict(
                "discipline",
                spec["discipline_name"],
                code=spec.get("discipline_code"),
            )
            # 章节节点
            chapter_defs = spec.get("chapters") or {}
            chapters = {}
            if chapter_defs:
                for ch_code, ch_name in chapter_defs.items():
                    chapters[ch_code] = self._ensure_dict(
                        "chapter", ch_name, code=ch_code, parent=discipline
                    )
            else:
                chapters["__DEFAULT__"] = self._ensure_dict(
                    "chapter",
                    "通用章节",
                    code=f"{spec.get('discipline_code', sheet.title[:1])}0",
                    parent=discipline,
                )

            # 表头映射
            header_row = self.header_row or 1
            header_map = {}
            for c in range(1, sheet.max_column + 1):
                title = self._str(sheet, header_row, c)
                if title:
                    header_map[title] = c

            def col(*names):
                for n in names:
                    if n in header_map:
                        return header_map[n]
                return None

            c_code = col("定额名称", "定额编号")
            c_name = col("项目名称")
            c_unit = col("单位")
            c_total = col("综合单价")
            c_direct = col("直接费")
            c_labor = col("人工费")
            c_material = col("材料费")
            c_machine = col("机械费")
            c_rate = col("费率", "机械费率")
            c_misc = col("综合费")
            c_work = col("工作内容")

            if not (c_code and c_name and c_unit):
                raise UserError(f"工作表“{sheet.title}”缺少必需列（定额编号/项目名称/单位）。")

            created = updated = 0
            for r in range(header_row + 1, sheet.max_row + 1):
                quota_code = self._str(sheet, r, c_code)
                name = self._str(sheet, r, c_name)
                if not (quota_code or name):
                    continue

                unit_raw = self._str(sheet, r, c_unit)
                price_total = self._val(sheet, r, c_total)
                price_direct = self._val(sheet, r, c_direct)
                price_labor = self._val(sheet, r, c_labor)
                price_material = self._val(sheet, r, c_material)
                price_machine = self._val(sheet, r, c_machine)
                rate_misc = self._val(sheet, r, c_rate)
                amount_misc = self._val(sheet, r, c_misc)
                work_content = self._str(sheet, r, c_work) if c_work else ""

                # 章节归属
                if chapter_defs:
                    prefix = (quota_code or "")[:2]
                    chapter = chapters.get(prefix)
                    if not chapter:
                        chapter = chapters.setdefault(
                            "__UNCLASSIFIED__",
                            self._ensure_dict("chapter", "未分类章节", parent=discipline),
                        )
                else:
                    chapter = chapters["__DEFAULT__"]

                # 单位匹配
                uom = False
                if unit_raw:
                    uom = Uom.search([("name", "=", unit_raw)], limit=1)

                # 查找/创建子目（章节+code 唯一）
                domain = [("type", "=", "sub_item"), ("parent_id", "=", chapter.id)]
                if quota_code:
                    domain.append(("code", "=", quota_code))
                else:
                    domain.append(("name", "=", name))
                sub = Dic.search(domain, limit=1)

                vals = {
                    "type": "sub_item",
                    "name": name,
                    "code": quota_code or False,
                    "quota_code": quota_code or False,
                    "parent_id": chapter.id,
                    "uom_id": uom.id if uom else False,
                    "price_total": float(price_total or 0.0),
                    "price_direct": float(price_direct or 0.0),
                    "price_labor": float(price_labor or 0.0),
                    "price_material": float(price_material or 0.0),
                    "price_machine": float(price_machine or 0.0),
                    "rate_misc": float(rate_misc or 0.0),
                    "amount_misc": float(amount_misc or 0.0),
                    "work_content": work_content or False,
                    "raw_line": json.dumps(
                        {
                            "sheet": sheet.title,
                            "row_index": r,
                            "quota_code": quota_code,
                            "name": name,
                            "unit": unit_raw,
                            "price_total": price_total,
                            "price_direct": price_direct,
                            "price_labor": price_labor,
                            "price_material": price_material,
                            "price_machine": price_machine,
                            "rate_misc": rate_misc,
                            "amount_misc": amount_misc,
                            "work_content": work_content,
                        },
                        ensure_ascii=False,
                    ),
                }

                if sub:
                    sub.write(vals)
                    updated += 1
                else:
                    Dic.create(vals)
                    created += 1

            total_created += created
            total_updated += updated
            log_lines.append(
                f"- 工作表 {sheet.title}: 创建 {created} 条，更新 {updated} 条"
            )

        self.log = "\n".join(log_lines)
        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": "定额导入完成",
                "message": f"创建 {total_created} 条，更新 {total_updated} 条。\n" + "\n".join(log_lines),
                "type": "success",
                "sticky": False,
            },
        }
